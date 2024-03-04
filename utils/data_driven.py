import openpyxl


class DatosConfig:
    def __init__(self, ambiente):
        self.ambiente = ambiente

    def obtener_datos_xlsx(self, nombre_hoja):
        """
        Método para obtener los datos de la hoja de excel\n
        Ejemplo de uso:\n
        datos = DatosConfig("qa")\n
        datos.obtener_datos("Sheet1")\n
        este método retorna una lista con los datos de la hoja de excel
        """
        # Cargar el archivo de excel
        rutaArchivo = f"data/datos_{self.ambiente}.xlsx"
        print(rutaArchivo)
        archivo = openpyxl.load_workbook(rutaArchivo)
        hoja = archivo[nombre_hoja]

        # Obtener los datos de la hoja
        datos = []
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column):
            fila = [celda.value for celda in fila]
            datos.append(fila)

        return datos
